#Import the required module
import requests
from bs4 import BeautifulSoup
import csv
import datetime
import time
import openpyxl
import os

#Define the headers for the web requests
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"}

#Define a function to get the stock symbols from a folder
def get_stock_symbols(folder):
    #Create an empty list to store the results
    symbols = []
    #loop through each file in the folder
    for file in os.listdir(folder):
        #Check if the file is an Excel file
        if file.endswith(".xlsx"):
            #Extract the stock symbol from the file name
            symbol = file.split(".")[0]
            #Append the symbol to the symbols list
            symbols.append(symbol)
    #Return the symbols list
    return symbols

folder = "C:\\Users\\mossa\\OneDrive\\Aktier\\"
symbols = get_stock_symbols(folder)

#Define a function to get the stock prices from Google
def get_stock_prices(symbols):
    #Create an empty dictionary to store the results
    results = {}
    #Open the csv file in append mode
    with open('stock_prices.csv', 'a', newline='', encoding='UTF8') as f:
        #Create a csv writer object
        writer = csv.writer(f)
        #Write the header only if the file is empty
        if f.tell() == 0:
            header = ['Stock', 'Price', 'Date']
            writer.writerow(header)
        #Loop through each symbol
        for symbol in symbols:
            #Construct the url for the web request
            url = f"https://www.google.com/search?q={symbol}+stock+price"
            #Make the web request and parse the html response
            res = requests.get(url, headers=headers)
            soup = BeautifulSoup(res.text, "html.parser")
            #Try to extract the price from the html element with class "wT3VGc"
            try:
                price = soup.select_one(".wT3VGc").text
                results[symbol] = price
            #If the price cannot be extracted, set it to an error message
            except:
                results[symbol] = "Failed to retrieve stock price"
            #Write a new row of data for each stock
            data = [symbol, price, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            writer.writerow(data)

    #Return the results dictionary
    return results


#Define a function to write the stock prices to separate Excel files
def write_to_excel(prices):
    #Loop through each symbol and price in the prices dictionary
    for symbol, price in prices.items():
        #Create a file name for each Excel file based on the symbol
        file_name = f"C:\\Users\\User\\bioinformatics\\MyProjects\\scraping_stock_prizes\\Aktier\\{symbol}.xlsx"
        #Try to open an existing Excel file with that name or create a new one if it does not exist
        try:
            wb = openpyxl.load_workbook(file_name)
        except:
            wb = openpyxl.Workbook()
        #Get the second work sheet in the workbook by name
        ws = wb["KÖP-LONG"]
        #Find the next empty row in column A by looping through each cell
        row = 11
        while ws.cell(row=row, column=1).value != None:
            row += 1
        #Write the date and price to columns A and B respectively in that row
        ws.cell(row=row, column=1).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row, column=2).value = price.replace(",", ".")
        #Save and close the workbook
        wb.save(file_name)
        wb.close()

#Run the functions once a day inside a while loop
while True:
    prices = get_stock_prices(symbols)
    write_to_excel(prices)
    time.sleep(86400)

